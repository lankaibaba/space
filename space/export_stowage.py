"""
配载单明细导出 & 零担表融合脚本
- 导出1：昨天配载日期（created_date + exe_pur_order_b.order_date）的订单
- 导出2：昨天签收时间（receive_time + exe_pur_order_b.order_date）的订单

流程：
1. 登录获取token
2. 调用 async-export API 提交导出任务
3. 轮询 queued-task/my/{task_id} 等待任务完成
4. 从任务结果中获取 fileKey
5. 获取临时下载授权码
6. 下载标准格式的 xlsx 文件
7. 融合零担表：
   a. 筛选配载表BS列五省数据，B-S列追加到零担表34月sheet末尾（F列订单状态不填充）
   b. 匹配签收表更新零担表F列空白行
"""

import requests
import base64
import json
import time
import sys
import os
import copy
from datetime import datetime, timedelta, timezone
from Crypto.PublicKey import RSA
from Crypto.Cipher import PKCS1_v1_5
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

# ====================== 配置 ======================
MY_ACCOUNT = "V0014052"
MY_PASSWORD = "Xby1234567"
PUBLIC_KEY = "MIGfMA0GCSqGSIb3DQEBAQUAA4GNADCBiQKBgQCctQTweXAiaQ3ct5bhj6nyisOQiGmgC/hUdK+QO9I9DudcQSUMxIXvMtpiogB9RWkAUC4b86x7SiGD6aCp7PbTspd5fLf8F6LUIj/BtmktQq7JNsShjAWBxCkE49HIIvPvl9rt8lO7MkgS2vUT04tEYeu/62ltOc3BljJXoPC4pQIDAQAB"

LOGIN_URL = "https://sdm.etransfar.com/jbl/api/login/?_allow_anonymous=true"
ASYNC_EXPORT_URL = "https://sdm.etransfar.com/jbl/api/module-data/receive_management/async-export"
QUEUED_TASK_URL = "https://sdm.etransfar.com/jbl/api/queued-task/my/{}"
AUTH_CODE_URL = "https://sdm.etransfar.com/jbl/api/file/get-temporary-auth-code?key={}"
DOWNLOAD_URL = "https://sdm.etransfar.com/jbl/api/file/download/{}?authCode={}"
BASE_URL = "https://sdm.etransfar.com"

# 下载文件保存目录（脚本所在目录）
SAVE_DIR = os.path.dirname(os.path.abspath(__file__))

# 零担表路径
LINGDAN_TABLE_PATH = r"C:\Users\Administrator\Desktop\26年零担表.xlsx"

# 需要筛选的省份（BS列"收货区域省份"中包含这些关键词的行）
FILTER_PROVINCES = ["湖北", "湖南", "河北", "新疆", "安徽"]

# 导出列顺序（67列，与网页端导出一致）
EXPORT_SORT_FIELDS = [
    "source_order_no", "exe_pur_order_b.order_date", "delivery_date", "receive_time",
    "stowage_all_weight", "signed_weight", "status_dk", "receive_name", "detailed_address",
    "receiver_name", "receiver_phone", "salesman_name", "receive_region_code", "material_name",
    "packaging_num", "plate_no", "carrier_name", "carrier_phone", "no", "order_no",
    "type_code", "sub_type_code", "expected_arrival_date", "k_contract_line_a.network",
    "exe_pur_order_b.pur_org_code", "created_date", "start_time", "receiver",
    "receive_registrant", "exe_pur_order_b.customer", "exe_pur_order_b.total_weight",
    "shipping_weight", "refund_weight", "exceed_weight", "send_address", "send_name",
    "exe_delivery_note_b.attachment", "sign_attachment", "arrive_attachment",
    "exe_pur_order_b.purveyor", "receipt_time", "exe_pur_order_b.customer_no",
    "receive_prescription", "receive_source", "exe_delivery_note_b.is_borrow_sold",
    "exe_delivery_note_b.ship_no", "exe_delivery_note_b.container_no", "order_type_code",
    "exe_pur_order_b.send_storage_code", "audit_check", "exe_delivery_note_b.yz_flag",
    "exe_pur_order_b.receiver_name", "exe_pur_order_b.receiver_phone", "posting_date",
    "sign_status", "exe_pur_order_b.customer_group", "send_region_code", "is_sent_out",
    "is_post", "exe_pur_order_b.mat_desc", "exe_pur_order_b.change_label_flag",
    "sent_out_date", "the_way_flag", "exe_pur_order_b.sale_org",
    "exe_pur_order_b.customer_grade", "logistics_time", "customer_prescription"
]


# ====================== 工具函数 ======================
def rsa_encrypt(password):
    pub_key = f"-----BEGIN PUBLIC KEY-----\n{PUBLIC_KEY}\n-----END PUBLIC KEY-----"
    rsa_key = RSA.importKey(pub_key)
    cipher = PKCS1_v1_5.new(rsa_key)
    encrypted = cipher.encrypt(password.encode())
    return base64.b64encode(encrypted).decode()


def login():
    """登录获取token"""
    payload = {
        "name": MY_ACCOUNT,
        "password": rsa_encrypt(MY_PASSWORD),
        "rememberMe": True,
        "imageCode": None,
        "loginBindingParameters": {}
    }
    headers = {
        "Content-Type": "application/json;charset=UTF-8",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Referer": "https://sdm.etransfar.com/jbl/"
    }
    resp = requests.post(LOGIN_URL, json=payload, headers=headers, timeout=15)
    data = resp.json()
    if data.get("status") == "login":
        print(f"[OK] 登录成功")
        return data["token"]
    print(f"[FAIL] 登录失败: {data}")
    return None


def get_auth_headers(token):
    """获取带认证的请求头"""
    return {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/147.0.0.0 Safari/537.36 Edg/147.0.0.0",
        "language": "zh_CN",
        "timezone": "UTC+0800",
        "Accept": "application/json, text/plain, */*",
        "Referer": "https://sdm.etransfar.com/",
        "Origin": "https://sdm.etransfar.com",
        "sec-ch-ua": '"Microsoft Edge";v="147", "Not.A/Brand";v="8", "Chromium";v="147"',
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": '"Windows"',
        "sec-fetch-dest": "empty",
        "sec-fetch-mode": "cors",
        "sec-fetch-site": "same-origin",
    }


def get_yesterday_utc_range():
    """获取昨天（北京时间）对应的UTC时间范围"""
    now_bj = datetime.now(timezone(timedelta(hours=8)))
    yesterday_bj = now_bj - timedelta(days=1)

    start_bj = yesterday_bj.replace(hour=0, minute=0, second=0, microsecond=0)
    end_bj = yesterday_bj.replace(hour=23, minute=59, second=59, microsecond=999000)

    start_utc = start_bj.astimezone(timezone.utc)
    end_utc = end_bj.astimezone(timezone.utc)

    return start_utc, end_utc


def build_export_payload(rules):
    """构建async-export请求体（与网页端格式完全一致）"""
    return {
        "direction": "DESC",
        "property": "id",
        "fromClientType": "pc",
        "number": 0,
        "sorts": [],
        "rules": rules,
        "size": 15,
        "specialConditions": [],
        "dynamicFormCode": "stowage_sign_receipt",
        "developmentSystemId": None,
        "debugFlag": False,
        "exportSortFields": EXPORT_SORT_FIELDS
    }


def submit_async_export(token, rules):
    """提交异步导出任务，返回task_id"""
    headers = get_auth_headers(token)
    payload = build_export_payload(rules)

    resp = requests.post(ASYNC_EXPORT_URL, json=payload, headers=headers, timeout=30)
    resp.raise_for_status()
    r = resp.json()

    if r.get("status") != "success":
        print(f"[FAIL] async-export 失败: {r.get('message', r)}")
        return None

    task_id = r.get("data", {}).get("id", "")
    task_code = r.get("data", {}).get("code", "")
    print(f"[OK] 导出任务已提交: task_id={task_id}, code={task_code}")
    return task_id


def poll_task_until_complete(token, task_id, max_wait=120):
    """轮询任务状态直到完成，返回outputParam中的fileKey"""
    headers = get_auth_headers(token)
    poll_url = QUEUED_TASK_URL.format(task_id)

    start_time = time.time()
    while time.time() - start_time < max_wait:
        resp = requests.get(poll_url, headers=headers, timeout=15)
        if resp.status_code != 200:
            print(f"[WARN] 轮询失败 status={resp.status_code}, {resp.text[:200]}")
            time.sleep(3)
            continue

        task_data = resp.json()
        status = task_data.get("statusEk", "")

        if status == "SUCCEED":
            # 从outputParam.content中提取fileKey
            output_param = task_data.get("outputParam", {})
            content_str = output_param.get("content", "")
            if not content_str:
                print("[FAIL] 任务成功但outputParam.content为空")
                return None

            content = json.loads(content_str)
            attachment = content.get("attachment", {})
            attach_files = attachment.get("attachFile", [])

            if not attach_files:
                print("[FAIL] 任务成功但没有附件")
                return None

            file_info = attach_files[0]
            file_key = file_info.get("key", "")
            file_name = file_info.get("name", "")
            file_size = file_info.get("fileSize", "")

            print(f"[OK] 任务完成: {file_name} ({file_size})")
            return file_key

        elif status in ("FAILED", "ERROR"):
            print(f"[FAIL] 任务失败: status={status}")
            output = task_data.get("outputParam", {}).get("content", "")
            print(f"  错误信息: {output[:300]}")
            return None

        # 任务还在处理中
        elapsed = int(time.time() - start_time)
        print(f"  等待中... ({elapsed}s) status={status}")
        time.sleep(3)

    print(f"[FAIL] 任务超时（{max_wait}s）")
    return None


def get_download_auth_code(token, file_key):
    """获取文件临时下载授权码"""
    headers = get_auth_headers(token)
    url = AUTH_CODE_URL.format(file_key)

    resp = requests.get(url, headers=headers, timeout=15)
    resp.raise_for_status()
    r = resp.json()

    auth_code = r.get("temporaryAuthCode", "")
    if auth_code:
        print(f"[OK] 获取下载授权码成功")
        return auth_code

    print(f"[FAIL] 获取下载授权码失败: {r}")
    return None


def download_file(token, file_key, auth_code, save_path):
    """下载导出的xlsx文件"""
    headers = get_auth_headers(token)
    url = DOWNLOAD_URL.format(file_key, auth_code)

    resp = requests.get(url, headers=headers, timeout=60)
    if resp.status_code != 200:
        print(f"[FAIL] 下载失败: status={resp.status_code}")
        return False

    content_type = resp.headers.get("Content-Type", "")
    if "force-download" not in content_type and "octet" not in content_type and "xlsx" not in content_type:
        # 可能返回了错误JSON
        if len(resp.content) < 1000:
            print(f"[FAIL] 下载可能失败: {resp.text[:200]}")
            return False

    # 如果文件被占用，先保存到临时文件
    actual_path = save_path
    try:
        with open(save_path, "wb") as f:
            f.write(resp.content)
    except PermissionError:
        base, ext = os.path.splitext(save_path)
        actual_path = f"{base}_new{ext}"
        with open(actual_path, "wb") as f:
            f.write(resp.content)
        print(f"[WARN] 原文件被占用，已保存到: {actual_path}")

    file_size = os.path.getsize(actual_path)
    print(f"[OK] 文件已保存: {actual_path} ({file_size} bytes)")
    return True


def export_and_download(token, rules, filename):
    """完整的导出流程：提交任务 → 轮询 → 获取授权码 → 下载"""
    save_path = os.path.join(SAVE_DIR, filename)

    # Step 1: 提交导出任务
    task_id = submit_async_export(token, rules)
    if not task_id:
        return False

    # Step 2: 轮询直到完成
    file_key = poll_task_until_complete(token, task_id)
    if not file_key:
        return False

    # Step 3: 获取下载授权码
    auth_code = get_download_auth_code(token, file_key)
    if not auth_code:
        return False

    # Step 4: 下载文件
    success = download_file(token, file_key, auth_code, save_path)
    return success


# ====================== 零担表融合函数 ======================

# 配载表列 -> 零担表列 的映射
# 配载表B(来源单创建日期) -> 零担表B(创建时间)
# 配载表C(要求到货时间) -> 零担表C(要求到货时间)
# 配载表A(来源单号) -> 零担表D(来源单号)
# 配载表E(配载总重量) -> 零担表E(配载总重量)
# 配载表G(配载单状态) -> 零担表F(订单状态)
# 零担表G(回单状态) -> 留空
# 配载表H(收货方名称) -> 零担表H(收货方名称)
# 配载表I(收货地址) -> 零担表I(收货地址)
# 配载表J(收货联系人) -> 零担表J(收货联系人)
# 配载表K(收货联系人电话) -> 零担表K(收货电话)
# 配载表L(业务员姓名) -> 零担表L(业务员姓名)
# 配载表P(车牌号) -> 零担表M(车牌)
# 配载表Q(司机名称) -> 零担表N(司机)
# 配载表R(司机联系方式) -> 零担表O(司机电话)
# 配载表S(配载单号) -> 零担表P(配载单号)
# 配载表BS(收货区域省份) -> 零担表Q(省)
# 配载表BT(收货区域城市) -> 零担表R(市)
# 配载表BU(收货区域区县) -> 零担表S(区)
# 零担表T(备注) -> 留空
STOWAGE_TO_LINGDAN_MAP = [
    # (零担表列号, 配载表列号)
    (2, 2),   # B -> B: 来源单创建日期 -> 创建时间
    (3, 3),   # C -> C: 要求到货时间
    (4, 1),   # D -> A: 来源单号
    (5, 5),   # E -> E: 配载总重量
    # F(订单状态) 不填充
    # G(回单状态) 留空
    (8, 8),   # H -> H: 收货方名称
    (9, 9),   # I -> I: 收货地址
    (10, 10), # J -> J: 收货联系人
    (11, 11), # K -> K: 收货联系人电话
    (12, 12), # L -> L: 业务员姓名
    (13, 16), # M -> P: 车牌 -> 车牌号
    (14, 17), # N -> Q: 司机 -> 司机名称
    (15, 18), # O -> R: 司机电话 -> 司机联系方式
    (16, 19), # P -> S: 配载单号
    (17, 71), # Q -> BS: 省 -> 收货区域省份
    (18, 72), # R -> BT: 市 -> 收货区域城市
    (19, 73), # S -> BU: 区 -> 收货区域区县
]


def format_time_display(val):
    """将UTC时间字符串转为北京时间显示格式"""
    if not val:
        return None
    try:
        dt_utc = datetime.fromisoformat(str(val).replace("Z", "+00:00"))
        dt_bj = dt_utc.astimezone(timezone(timedelta(hours=8)))
        return dt_bj.strftime("%Y-%m-%d %H:%M")
    except:
        return val


def merge_stowage_to_lingdan(stowage_path, lingdan_path, sign_path=None):
    """
    步骤1：筛选配载表BS列五省数据，B-S列追加到零担表34月sheet末尾
    步骤2：匹配签收表更新零担表F列空白行
    """
    print("\n" + "=" * 60)
    print("   零担表融合处理")
    print("=" * 60)

    if not os.path.exists(stowage_path):
        print(f"[FAIL] 配载表不存在: {stowage_path}")
        return False
    if not os.path.exists(lingdan_path):
        print(f"[FAIL] 零担表不存在: {lingdan_path}")
        return False

    # 读取配载表
    print("\n--- 读取配载表 ---")
    wb_stowage = openpyxl.load_workbook(stowage_path)
    ws_stowage = wb_stowage.active

    # 筛选BS列（收货区域省份）包含五省的数据
    bs_col = 71  # BS列
    filtered_rows = []
    for r in range(2, ws_stowage.max_row + 1):
        province = ws_stowage.cell(r, bs_col).value
        if province and any(p in str(province) for p in FILTER_PROVINCES):
            filtered_rows.append(r)

    print(f"[OK] 配载表共 {ws_stowage.max_row - 1} 行，筛选出 {len(filtered_rows)} 行（五省数据）")

    # 统计各省数据
    province_count = {}
    for r in filtered_rows:
        province = ws_stowage.cell(r, bs_col).value
        for p in FILTER_PROVINCES:
            if p in str(province):
                province_count[p] = province_count.get(p, 0) + 1
                break
    for p, cnt in province_count.items():
        print(f"  {p}: {cnt} 行")

    # 读取零担表
    print("\n--- 读取零担表 ---")
    wb_lingdan = openpyxl.load_workbook(lingdan_path)
    ws_lingdan = wb_lingdan['34月']
    original_max_row = ws_lingdan.max_row
    print(f"[OK] 零担表34月 sheet 当前 {original_max_row} 行")

    # 获取格式模板（从最后一行复制格式，包含行高等参数）
    template_row = original_max_row
    template_cell_formats = {}
    for c in range(1, 21):  # A-T
        cell = ws_lingdan.cell(template_row, c)
        template_cell_formats[c] = {
            'font': copy.copy(cell.font),
            'fill': copy.copy(cell.fill),
            'border': copy.copy(cell.border),
            'alignment': copy.copy(cell.alignment),
            'number_format': cell.number_format,
        }
    # 获取模板行高
    template_row_height = ws_lingdan.row_dimensions[template_row].height

    # 步骤1：将筛选后的数据追加到零担表末尾（去重）
    print("\n--- 步骤1：追加配载数据到零担表 ---")

    # 收集零担表中已有的来源单号（D列），用于去重
    existing_order_nos = set()
    for r in range(2, original_max_row + 1):
        val = ws_lingdan.cell(r, 4).value  # D列: 来源单号
        if val:
            existing_order_nos.add(str(val).strip())

    # 过滤掉已存在的单号
    new_filtered_rows = []
    skipped_count = 0
    for stowage_row in filtered_rows:
        order_no = ws_stowage.cell(stowage_row, 1).value  # A列: 来源单号
        if order_no and str(order_no).strip() in existing_order_nos:
            skipped_count += 1
        else:
            new_filtered_rows.append(stowage_row)
            # 加入集合防止同一次追加中重复
            if order_no:
                existing_order_nos.add(str(order_no).strip())

    if skipped_count > 0:
        print(f"[OK] 去重：跳过 {skipped_count} 条已存在的数据")

    if not new_filtered_rows:
        print("[OK] 没有新数据需要追加")
        wb_stowage.close()
        # 仍继续步骤2（签收匹配）
    else:
        print(f"[OK] 待追加 {len(new_filtered_rows)} 条新数据")

    new_row_start = original_max_row + 1
    for i, stowage_row in enumerate(new_filtered_rows):
        lingdan_row = new_row_start + i

        # A列：序号（后续统一重排，先留空）
        ws_lingdan.cell(lingdan_row, 1).value = None

        # B-S列：按映射关系写入
        for lingdan_col, stowage_col in STOWAGE_TO_LINGDAN_MAP:
            val = ws_stowage.cell(stowage_row, stowage_col).value
            # 时间字段格式化
            if stowage_col in (2, 3):  # B列(来源单创建日期), C列(要求到货时间)
                val = format_time_display(val)
            ws_lingdan.cell(lingdan_row, lingdan_col).value = val

        # G列(回单状态)留空
        ws_lingdan.cell(lingdan_row, 7).value = None
        # T列(备注)留空
        ws_lingdan.cell(lingdan_row, 20).value = None

        # 应用格式（格式刷）
        for c in range(1, 21):
            cell = ws_lingdan.cell(lingdan_row, c)
            fmt = template_cell_formats[c]
            cell.font = fmt['font']
            cell.fill = fmt['fill']
            cell.border = fmt['border']
            cell.alignment = fmt['alignment']
            cell.number_format = fmt['number_format']

        # 复制行高
        if template_row_height is not None:
            ws_lingdan.row_dimensions[lingdan_row].height = template_row_height

    # 重排A列序号
    print("[OK] 重排序号...")
    for r in range(2, new_row_start + len(new_filtered_rows)):
        ws_lingdan.cell(r, 1).value = r - 1

    # 保存配载表（释放）
    wb_stowage.close()

    print(f"[OK] 已追加 {len(new_filtered_rows)} 行数据（行 {new_row_start}-{new_row_start + len(new_filtered_rows) - 1}）")

    # 步骤2：匹配签收表更新零担表F列空白行
    # 先保存零担表当前状态（步骤1的结果），再读取签收表
    print("\n--- 步骤2：匹配签收表更新F列 ---")

    # 读取签收表（优先使用传入的路径，否则按旧格式兼容查找）
    if sign_path is None:
        yesterday_str = datetime.now(timezone(timedelta(hours=8))).replace(
            hour=0, minute=0, second=0, microsecond=0
        ) - timedelta(days=1)
        yesterday_str_fmt = yesterday_str.strftime("%Y-%m-%d")
        sign_path = os.path.join(SAVE_DIR, f"昨天签收时间订单_{yesterday_str_fmt}.xlsx")

    if not os.path.exists(sign_path):
        # 尝试其他可能的名字
        alt_sign_path = os.path.join(SAVE_DIR, f"昨天签收时间订单_{yesterday_str_fmt}_new.xlsx")
        if os.path.exists(alt_sign_path):
            sign_path = alt_sign_path
        else:
            print(f"[WARN] 签收表不存在: {sign_path}，跳过步骤2")
            # 仍然保存零担表
            wb_lingdan.save(lingdan_path)
            wb_lingdan.close()
            return True

    wb_sign = openpyxl.load_workbook(sign_path)
    ws_sign = wb_sign.active

    # 构建签收表查找字典：A列(来源单号) -> AR列(签收来源)
    sign_dict = {}
    for r in range(2, ws_sign.max_row + 1):
        order_no = ws_sign.cell(r, 1).value  # A列: 来源单号
        sign_source = ws_sign.cell(r, 44).value  # AR列: 签收来源
        if order_no:
            sign_dict[str(order_no).strip()] = str(sign_source).strip() if sign_source else ""

    print(f"[OK] 签收表共 {len(sign_dict)} 条记录")

    # 遍历零担表，找F列(订单状态)为空白的行
    updated_count = 0
    not_found_count = 0
    new_max_row = new_row_start + len(filtered_rows)

    for r in range(2, new_max_row):
        f_val = ws_lingdan.cell(r, 6).value  # F列: 订单状态
        if f_val is None or str(f_val).strip() == "":
            d_val = ws_lingdan.cell(r, 4).value  # D列: 来源单号
            if d_val:
                order_no = str(d_val).strip()
                if order_no in sign_dict:
                    sign_source = sign_dict[order_no]
                    if sign_source == "PC":
                        ws_lingdan.cell(r, 6).value = "PC已签收"
                    elif sign_source == "微信":
                        ws_lingdan.cell(r, 6).value = "已签收"
                    else:
                        ws_lingdan.cell(r, 6).value = sign_source if sign_source else ""
                    updated_count += 1
                else:
                    not_found_count += 1

    wb_sign.close()

    print(f"[OK] F列更新完成: 更新 {updated_count} 行, 未匹配 {not_found_count} 行")

    # 保存零担表
    try:
        wb_lingdan.save(lingdan_path)
        print(f"[OK] 零担表已保存: {lingdan_path}")
    except PermissionError:
        base, ext = os.path.splitext(lingdan_path)
        fallback_path = f"{base}_new{ext}"
        wb_lingdan.save(fallback_path)
        print(f"[WARN] 原文件被占用，零担表已保存到: {fallback_path}")
    wb_lingdan.close()

    return True


def main():
    print("=" * 60)
    print("   配载单明细导出工具（系统标准格式）")
    print("=" * 60)

    # 输入天数：1=昨天，2=昨天+前天，3=昨天+前天+大前天...
    while True:
        try:
            days_input = input("\n请输入导出天数（1=仅昨天，2=昨天+前天，以此类推，默认1）：").strip()
            if not days_input:
                days_input = 1
            else:
                days_input = int(days_input)
            if days_input < 1:
                print("请输入大于0的整数")
                continue
            break
        except ValueError:
            print("请输入有效的整数")

    # 计算日期范围（北京时间）
    now_bj = datetime.now(timezone(timedelta(hours=8)))
    yesterday_bj = now_bj - timedelta(days=1)
    yesterday_str = yesterday_bj.strftime("%Y-%m-%d")

    start_utc, end_utc = get_yesterday_utc_range()
    start_utc = start_utc - timedelta(days=days_input - 1)  # 起始点往前推

    start_date_bj = start_utc.astimezone(timezone(timedelta(hours=8)))
    end_date_bj = end_utc.astimezone(timezone(timedelta(hours=8)))

    print(f"\n当前北京时间: {now_bj.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"导出范围: {start_date_bj.strftime('%Y-%m-%d')} ~ {end_date_bj.strftime('%Y-%m-%d')}（共{days_input}天）")
    print(f"UTC范围: {start_utc.strftime('%Y-%m-%dT%H:%M:%S.000Z')} ~ {end_utc.strftime('%Y-%m-%dT%H:%M:%S.999Z')}")

    # 登录
    token = login()
    if not token:
        print("登录失败，退出")
        sys.exit(1)

    # 昨天的时间范围
    yesterday_values = [
        start_utc.strftime("%Y-%m-%dT%H:%M:%S.000Z"),
        end_utc.strftime("%Y-%m-%dT%H:%M:%S.999Z")
    ]

    # 近一个月的时间范围
    one_month_ago_bj = now_bj - timedelta(days=31)
    one_month_ago_start = one_month_ago_bj.replace(hour=0, minute=0, second=0, microsecond=0).astimezone(timezone.utc)
    today_end = now_bj.replace(hour=23, minute=59, second=59, microsecond=999000).astimezone(timezone.utc)
    one_month_values = [
        one_month_ago_start.strftime("%Y-%m-%dT%H:%M:%S.000Z"),
        today_end.strftime("%Y-%m-%dT%H:%M:%S.999Z")
    ]

    # 文件名中的日期范围
    date_range_str = f"{start_date_bj.strftime('%Y-%m-%d')}_{end_date_bj.strftime('%Y-%m-%d')}"

    results = {}

    # ==================== 导出1：昨天配载日期的订单 ====================
    print("\n" + "=" * 60)
    print("导出1: 配载日期（created_date + order_date）的订单")
    print("=" * 60)

    rules1 = [
        {"field": "created_date", "option": "BTS", "values": yesterday_values},
        {"field": "exe_pur_order_b.order_date", "option": "BTS", "values": one_month_values}
    ]

    filename1 = f"配载日期订单_{date_range_str}.xlsx"
    success1 = export_and_download(token, rules1, filename1)
    results[filename1] = success1

    # ==================== 导出2：昨天签收时间的订单 ====================
    print("\n" + "=" * 60)
    print("导出2: 昨天签收时间（receive_time + order_date）的订单")
    print("=" * 60)

    rules2 = [
        {"field": "receive_time", "option": "BTS", "values": yesterday_values},
        {"field": "exe_pur_order_b.order_date", "option": "BTS", "values": one_month_values}
    ]

    filename2 = f"签收时间订单_{date_range_str}.xlsx"
    success2 = export_and_download(token, rules2, filename2)
    results[filename2] = success2

    # ==================== 步骤3：零担表融合 ====================
    if success1:
        stowage_path = os.path.join(SAVE_DIR, filename1)
        sign_file_path = os.path.join(SAVE_DIR, filename2) if success2 else None
        merge_result = merge_stowage_to_lingdan(stowage_path, LINGDAN_TABLE_PATH, sign_file_path)
        results["零担表融合"] = merge_result
    else:
        results["零担表融合"] = False
        print("\n[WARN] 配载表导出失败，跳过零担表融合")

    # ==================== 结果汇总 ====================
    print("\n" + "=" * 60)
    print("   导出结果汇总")
    print("=" * 60)
    for name, success in results.items():
        status = "成功" if success else "失败"
        print(f"  {name}: {status}")

    if not all(results.values()):
        print("\n部分导出失败，请登录系统手动下载:")
        print("  https://sdm.etransfar.com/jbl/")
    else:
        print(f"\n所有文件已下载到: {SAVE_DIR}")


if __name__ == "__main__":
    main()
